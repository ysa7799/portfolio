"""
Operations KPI Dashboard
-------------------------
Simulates 12 months of security-operations incident data (the kind logged daily
at a G4S-style site) and produces a clean Excel workbook with:
  - Raw data sheet
  - Monthly KPI summary (incident volume, avg response time, category mix)
  - A native Excel chart (no external image dependency)

Run: python3 generate_dashboard.py
Output: data/operations_kpi_dashboard.xlsx
"""
import random
from datetime import date, timedelta
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

random.seed(42)

CATEGORIES = ["Access Control", "Alarm Response", "Patrol Finding", "Equipment Fault", "Visitor Incident"]
START = date(2025, 1, 1)
DAYS = 365

rows = []
for i in range(DAYS):
    d = START + timedelta(days=i)
    # 0-6 incidents/day, seasonal bump in summer months
    seasonal = 1.4 if d.month in (6, 7, 8) else 1.0
    n_incidents = max(0, int(random.gauss(3, 1.5) * seasonal))
    for _ in range(n_incidents):
        cat = random.choices(CATEGORIES, weights=[30, 20, 25, 10, 15])[0]
        response_min = max(1, round(random.gauss(12 if cat == "Alarm Response" else 22, 6), 1))
        rows.append({"date": d, "category": cat, "response_time_min": response_min})

df = pd.DataFrame(rows)
df["month"] = pd.to_datetime(df["date"]).dt.strftime("%Y-%m")

# --- Monthly KPI summary ---
monthly = df.groupby("month").agg(
    incident_volume=("category", "count"),
    avg_response_min=("response_time_min", "mean"),
).reset_index()
monthly["avg_response_min"] = monthly["avg_response_min"].round(1)

cat_by_month = df.pivot_table(index="month", columns="category", values="date", aggfunc="count", fill_value=0)

# --- Write workbook ---
wb = Workbook()

# Raw data sheet
ws_raw = wb.active
ws_raw.title = "Raw Data"
ws_raw.append(["date", "category", "response_time_min", "month"])
for _, r in df.iterrows():
    ws_raw.append([r["date"].isoformat(), r["category"], r["response_time_min"], r["month"]])

# KPI summary sheet
ws_kpi = wb.create_sheet("KPI Summary")
header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)

ws_kpi.append(["month", "incident_volume", "avg_response_min"])
for c in ws_kpi[1]:
    c.fill = header_fill
    c.font = header_font
for _, r in monthly.iterrows():
    ws_kpi.append([r["month"], int(r["incident_volume"]), r["avg_response_min"]])

for col in ("A", "B", "C"):
    ws_kpi.column_dimensions[col].width = 18

# Volume chart
chart1 = BarChart()
chart1.title = "Monthly Incident Volume"
chart1.y_axis.title = "Incidents"
chart1.x_axis.title = "Month"
data = Reference(ws_kpi, min_col=2, min_row=1, max_row=ws_kpi.max_row)
cats = Reference(ws_kpi, min_col=1, min_row=2, max_row=ws_kpi.max_row)
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
ws_kpi.add_chart(chart1, "E2")

# Response time trend
chart2 = LineChart()
chart2.title = "Avg Response Time (min)"
data2 = Reference(ws_kpi, min_col=3, min_row=1, max_row=ws_kpi.max_row)
chart2.add_data(data2, titles_from_data=True)
chart2.set_categories(cats)
ws_kpi.add_chart(chart2, "E18")

# Category breakdown sheet
ws_cat = wb.create_sheet("Category Trends")
ws_cat.append(["month"] + list(cat_by_month.columns))
for c in ws_cat[1]:
    c.fill = header_fill
    c.font = header_font
for month, row in cat_by_month.iterrows():
    ws_cat.append([month] + list(row.values))

for i, col in enumerate(ws_cat.columns, start=1):
    ws_cat.column_dimensions[get_column_letter(i)].width = 16

chart3 = BarChart()
chart3.type = "col"
chart3.grouping = "stacked"
chart3.overlap = 100
chart3.title = "Incident Category Mix by Month"
data3 = Reference(ws_cat, min_col=2, max_col=ws_cat.max_column, min_row=1, max_row=ws_cat.max_row)
cats3 = Reference(ws_cat, min_col=1, min_row=2, max_row=ws_cat.max_row)
chart3.add_data(data3, titles_from_data=True)
chart3.set_categories(cats3)
ws_cat.add_chart(chart3, "I2")

# Insights sheet
ws_notes = wb.create_sheet("Insights")
ws_notes["A1"] = "Key Insights"
ws_notes["A1"].font = Font(bold=True, size=14, color="1F3864")
insights = [
    f"Total incidents logged (12 months): {len(df)}",
    f"Overall average response time: {df['response_time_min'].mean():.1f} min",
    f"Highest-volume month: {monthly.loc[monthly['incident_volume'].idxmax(), 'month']} "
    f"({monthly['incident_volume'].max()} incidents)",
    "Alarm Response consistently shows the fastest average response time — expected, as it is the "
    "highest-priority category and typically has a dedicated response team.",
    "Incident volume rises ~30-40% in June-August, consistent with higher footfall/seasonal activity — "
    "recommend adjusting shift staffing for those months.",
]
for i, line in enumerate(insights, start=3):
    ws_notes[f"A{i}"] = f"- {line}"
    ws_notes[f"A{i}"].alignment = Alignment(wrap_text=True)
ws_notes.column_dimensions["A"].width = 110

wb.save("data/operations_kpi_dashboard.xlsx")
print(f"Wrote data/operations_kpi_dashboard.xlsx — {len(df)} rows, {len(monthly)} months")
